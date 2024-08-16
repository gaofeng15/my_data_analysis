import os,re,pandas as pd,numpy as np,PySimpleGUI as sg
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Alignment,Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

# 不显示告警
pd.options.mode.chained_assignment = None

def set_excel_style(wb,output_path):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col)].width=25
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='宋体', size=12, bold=False, color='000000')
                cell.border = thin_border
    wb.save(output_path)

class Zhuangjizhibiao:
    # 设置表格表头
    def __init__(self):
        self.xitongbiao = pd.DataFrame()
        self.td1 = pd.Timedelta('1 day')
        # 预设要处理的表格
        self.FTTRjungongliangbiao=pd.DataFrame()
        self.chulibiao=pd.DataFrame()
        self.chulibiao_suixiao=pd.DataFrame()
        self.zongguidangbiao=pd.DataFrame()
        self.zaitubiao=pd.DataFrame()
        # 月累计指标sheet 3个表头
        self.YOUXIAOCHENGGONG_leiji = [['有效成功交付率', '当月装机、移机派单量', '竣工量', '撤单量']]
        self.DANGRICHENGGONG_leiji = [['当日成功交付率（月累计）', '当月装机、移机竣工量', '超24小时工单量']]
        self.FTTRJIAOFUCHENGGONG_leiji = [['FTTR交付成功率（月累计）', '当月FTTR装机、移机工单量', '撤回量', '竣工量']]
        self.jiakuanquanliang_rate = [['家宽全业务交付率','竣工量','派单量','撤单量','其他']]
        # 单日成功交付率sheet 2个表头
        self.DANGRICHENGGONG = [['日期', '24小时回单', '总回单', '当日成功交付率']]
        self.FTTRJIAOFUCHENGGONG = [['日期', 'FTTR24小时回单', '总回单', 'FTTR当日成功交付率']]
        self.workbook=Workbook()
    # 系统GUI框架及运行功能方法
    def run(self):
        # 系统布局
        layout = [[sg.Input(key='-IN-',disabled=True), sg.FileBrowse(target='-IN-', file_types=(('All Files', '*.xlsx'),))],
                  [sg.Input(key='-OUT-',disabled=True), sg.FileSaveAs(target='-OUT-', file_types=(('All Files', '*.xlsx'),))],
                  [sg.OK()]]
        window = sg.Window('指标统计', layout)
        while True:
            event, values = window.read()
            if event == None:
                break
            # 运行功能方法
            if event == 'OK':
                self.get_data(values['-IN-'])
                self.jiakuan_zhijia()
                self.month_targets()
                self.day_targets()
                self.zaitu_gongdan()
                self.FTTR_in_processing()
                self.suixiao(values['-IN-'])
                self.save_data(values['-OUT-'])
                sg.Popup('生成表格成功')
        window.close()
    # 获取所需数据
    def get_data(self,input_address):
        # 取消冻结行（使用openpyxl）
        wb = load_workbook(input_address)
        ws = wb.active
        ws.freeze_panes = None
        wb.save(input_address)
        # 筛选出需要的列，及装机、移机行（使用pandas）
        self.xitongbiao = pd.read_excel(input_address)
        try:
            self.chulibiao = self.xitongbiao[['工单编号','工单类型', '业务类型','小区名称','处理人', '工单状态', '建单时间', '回复时间']]
        except Exception:
            sg.Popup('文件有误')
        else:
            # 获取要处理的数据
            self.chulibiao = self.chulibiao[self.chulibiao.loc[:, '工单类型'] != '换']
            self.chulibiao['建单时间'] = pd.to_datetime(self.chulibiao['建单时间'])
            self.chulibiao['回复时间'] = pd.to_datetime(self.chulibiao['回复时间'])
            # 获取归档的数据
            self.zongguidangbiao = self.chulibiao[self.chulibiao.loc[:, '工单状态'] == '已竣工']
            self.zongguidangbiao = self.zongguidangbiao.assign(
                处理时长=self.zongguidangbiao['回复时间'] - self.zongguidangbiao['建单时间'])
    # 月累计指标sheet
    def month_targets(self):
            # 计算总派单,总归档量
        zongpaidan = self.chulibiao.shape[0]
        zongguidang = self.zongguidangbiao.shape[0]
            # 计算当月撤单量
        chedanbiao = self.chulibiao[self.chulibiao.loc[:, '工单状态'] == '已撤单']
        chedanliang = chedanbiao.shape[0]
            # 计算当月超时工单量
        chaoshigongdan = ((self.zongguidangbiao.loc[:, '处理时长'] > self.td1) * 1).sum()
            # 计算当月FTTR派单量
        FTTRgongdanbiao = self.chulibiao[self.chulibiao.loc[:, '业务类型'] == '智慧家庭(FTTR)']
        FTTRgongdan = FTTRgongdanbiao.shape[0]
            # 计算当月FTTR撤单量
        FTTRchedanliang = FTTRgongdanbiao[FTTRgongdanbiao.loc[:, '工单状态'] == '已撤单'].shape[0]
            # 计算当月FTTR竣工量
        self.FTTRjungongliangbiao = self.zongguidangbiao[self.zongguidangbiao.loc[:, '业务类型'] == '智慧家庭(FTTR)']
        FTTRjungongliang = self.FTTRjungongliangbiao.shape[0]

        # 计算有效成功交付率
        successfulDeliveryRate = '{:.2%}'.format(zongguidang / (zongpaidan - chedanliang))
        lst = [successfulDeliveryRate, zongpaidan, zongguidang, chedanliang]
        self.YOUXIAOCHENGGONG_leiji.append(lst)
        # 计算当月24小时内安装成功率
        oneDayInstallRate = '{:.2%}'.format((zongguidang - chaoshigongdan) / zongguidang)
        lst = [oneDayInstallRate, zongguidang, chaoshigongdan]
        self.DANGRICHENGGONG_leiji.append(lst)
        # 计算FTTR交付成功率（累计）
        FTTRsuccess_month = '{:.2%}'.format(FTTRjungongliang / (FTTRgongdan - FTTRchedanliang))
        lst = [FTTRsuccess_month, FTTRgongdan, FTTRchedanliang, FTTRjungongliang]
        self.FTTRJIAOFUCHENGGONG_leiji.append(lst)
        # 计算家客全业务交付率
        self.xitongbiao = self.xitongbiao[(self.xitongbiao['工单类型'] == '装') | (self.xitongbiao['工单类型'] == '移')]
        jiakuan_quanyewu_total = (
                (~ self.xitongbiao['业务类型'].str.contains('企业宽带')) & (self.xitongbiao['工单类型'] != '换')).sum()
        jiakuan_quanyewu_finish = (
                (~ self.xitongbiao['业务类型'].str.contains('企业宽带')) & (self.xitongbiao['工单状态'] == '已竣工') & (
                self.xitongbiao['工单类型'] != '换') & (self.xitongbiao['工单状态'] == '已竣工')).sum()
        jiaofu_rate = '{:.2%}'.format(jiakuan_quanyewu_finish / jiakuan_quanyewu_total)
        jiakuan_quanyewu_chedan = (
                (~ self.xitongbiao['业务类型'].str.contains('企业宽带')) & (self.xitongbiao['工单类型'] != '换') & (
                self.xitongbiao['工单状态'] == '已撤单')).sum()
        jiakuan_quanyewu_other=jiakuan_quanyewu_total-jiakuan_quanyewu_chedan-jiakuan_quanyewu_finish
        lst=[jiaofu_rate,jiakuan_quanyewu_finish,jiakuan_quanyewu_total,jiakuan_quanyewu_chedan,jiakuan_quanyewu_other]
        self.jiakuanquanliang_rate.append(lst)

        # 创建工作表
        nws=self.workbook.create_sheet('月累计指标')
        nws.column_dimensions['A'].width = 23.5
        nws.column_dimensions['B'].width = 26.4
        nws.column_dimensions['C'].width = 14.4
        nws.column_dimensions['D'].width = 14.4
        for row in self.YOUXIAOCHENGGONG_leiji:
            nws.append(row)
        for row in self.DANGRICHENGGONG_leiji:
            nws.append(row)
        for row in self.FTTRJIAOFUCHENGGONG_leiji:
            nws.append(row)
        for row in self.jiakuanquanliang_rate:
            nws.append(row)
    # 日交付成功率sheet
    def day_targets(self):
        # 计算每日安装成功的工单数
        for g in self.zongguidangbiao.groupby(self.zongguidangbiao.回复时间.dt.date):
            riqi = g[0]
            huidan_24 = np.sum(g[1]['处理时长'] < self.td1)
            huidan_total = g[1].shape[0]
            jiaofulv = '{:.2%}'.format(np.sum(g[1]['处理时长'] < self.td1) / g[1].shape[0])
            lst = [riqi, huidan_24, huidan_total, jiaofulv]
            self.DANGRICHENGGONG.append(lst)
        # 计算每日FTTR安装成功的工单数
        for g in self.FTTRjungongliangbiao.groupby(self.zongguidangbiao.回复时间.dt.date):
            riqi = g[0]
            huidan_24 = np.sum(g[1]['处理时长'] < self.td1)
            huidan_total = g[1].shape[0]
            jiaofulv = '{:.2%}'.format(np.sum(g[1]['处理时长'] < self.td1) / g[1].shape[0])
            lst = [riqi, huidan_24, huidan_total, jiaofulv]
            self.FTTRJIAOFUCHENGGONG.append(lst)
        nws=self.workbook.create_sheet('单日成功交付率')
        columns_to_adjust = ['A', 'B', 'C', 'D']
        for column in columns_to_adjust:
            nws.column_dimensions[column].width = 16.75
        for row in self.DANGRICHENGGONG:
            nws.append(row)
        for row in self.FTTRJIAOFUCHENGGONG:
            nws.append(row)
    # 在途工单量sheet
    def zaitu_gongdan(self):
        # 计算每个维护员的在途工单
        self.zaitubiao = self.chulibiao[(self.chulibiao['工单状态'] != '已撤单') & (self.chulibiao['工单状态'] != '已竣工')]
        self.zaitubiao = self.zaitubiao.loc[:, ['处理人', '工单类型','业务类型']].pivot_table(
            values='工单类型', index='处理人', columns='业务类型',aggfunc=np.count_nonzero, margins=True, margins_name='合计'
        ).rename(columns={'工单类型': '在途量'})
        # 为实现封装，将dataframe中的行索引、列索引、值整合成一个列表，然后保存到nws里
        zaitu_data=np.vstack((self.zaitubiao.columns.insert(0,'处理人'),np.column_stack((self.zaitubiao.index,self.zaitubiao.values)))).tolist()
        nws=self.workbook.create_sheet('维护员在途工单量')
        # 将数据写入工作表
        for lst in zaitu_data:
            nws.append(lst)
    # 在途FTTR明细sheet
    def FTTR_in_processing(self):
        FTTR_data=self.chulibiao.loc[:,['工单编号','工单类型','业务类型','小区名称','处理人','工单状态','建单时间']]
        FTTR_data['在途时长']=pd.Timestamp.now()-FTTR_data['建单时间']
        FTTR_data=FTTR_data[(FTTR_data['业务类型']=='智慧家庭(FTTR)')&(FTTR_data['工单状态']!='已撤单')&(FTTR_data['工单状态']!='已竣工')]
        FTTR_data_lst=np.vstack((FTTR_data.columns,FTTR_data.values)).tolist()
        nws = self.workbook.create_sheet('FTTR在途工单')
        columns_to_adjust = ['A', 'B', 'C', 'D','E','F','G','H']
        for column in columns_to_adjust:
            nws.column_dimensions[column].width = 19.5
        for lst in FTTR_data_lst:
            nws.append(lst)
    # 智家、家宽在途数据
    def jiakuan_zhijia(self):
        zaitubiao=self.chulibiao[(self.chulibiao.loc[:,'工单状态']!= '已撤单') & (self.chulibiao['工单状态'] != '已竣工')]
        zaitubiao['家宽']=zaitubiao.loc[:,'业务类型'].str.contains('家庭宽带')
        zaitubiao['智家']=(zaitubiao.loc[:, '业务类型'].str.contains('IPTV') \
              | zaitubiao.loc[:, '业务类型'].str.contains('爱家') \
              | zaitubiao.loc[:, '业务类型'].str.contains('路由') \
              | zaitubiao.loc[:, '业务类型'].str.contains('FTTR'))*1
        d=zaitubiao.loc[:, ['处理人', '家宽', '智家']].pivot_table(
            index='处理人',
            aggfunc={'家宽': 'sum',
                     '智家': 'sum'},
            margins=True,
            margins_name='合计'
        )
        jiakuan_zhijia_zaitu = np.vstack((d.columns.insert(0, '处理人'),
                            np.column_stack((d.index, d.values)))).tolist()
        nws = self.workbook.create_sheet('家宽、智家在途')
        # 将数据写入工作表
        for lst in jiakuan_zhijia_zaitu:
            nws.append(lst)
    def suixiao(self,input_address):
        xitongbiao = pd.read_excel(input_address)
        suixiaobiao = xitongbiao[xitongbiao['受理营业厅'].str.contains('随销').astype('bool')]
        suixiao_mingxi = suixiaobiao[((suixiaobiao['工单状态'] == '已竣工') | (suixiaobiao['工单状态'] == '已预约')) & (suixiaobiao['工单类型'] == '装')]
        # suixiao_mingxi =
        suixiao_data1 = suixiao_mingxi.assign(宽带=(suixiao_mingxi.loc[:, '业务类型'].str.contains('家庭宽带') | suixiao_mingxi.loc[:, '业务类型'].str.contains('企业宽带')) * 1)
        suixiao_data2 = suixiao_data1.assign(智家=suixiao_data1.loc[:, '业务类型'].str.contains('IPTV') \
                                                  | suixiao_data1.loc[:, '业务类型'].str.contains('爱家') \
                                                  | suixiao_data1.loc[:, '业务类型'].str.contains('路由') \
                                                  | suixiao_data1.loc[:, '业务类型'].str.contains('FTTR')
                                                  | suixiao_data1.loc[:, '业务类型'].str.contains('增强型')) * 1
        suixiao_data3 = suixiao_data2.loc[:, ['宽带', '智家', '工单状态', '受理营业员名字']]

        #
        self.chulibiao_suixiao = suixiao_data3.pivot_table(
            index='受理营业员名字',
            columns='工单状态',
            aggfunc='sum',
            margins=True,
            margins_name='总计'
        )


        # 为实现封装，将dataframe中的行索引、列索引、值整合成一个列表，然后保存到nws里
        suixiao_data=np.vstack((self.chulibiao_suixiao.columns.insert(0,'工单类型'),np.column_stack((self.chulibiao_suixiao.index,self.chulibiao_suixiao.values)))).tolist()

        nws=self.workbook.create_sheet('随销')
        lst1=[]
        lst2=[]
        for i,j in suixiao_data[0]:
            lst1.append(i)
            lst2.append(j)
        del suixiao_data[0]
        suixiao_data.insert(0,lst2)
        suixiao_data.insert(0, lst1)
        # 将数据写入工作表
        for lst in suixiao_data:
            nws.append(lst)


    # 保存数据到excel表
    def save_data(self,output_address):
        # 设置整体表格格式
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for sheet in self.workbook.sheetnames:
            ws = self.workbook[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(name='宋体', size=12, bold=False, color='000000')
                    cell.border = thin_border
        # 移除默认创建的sheet
        self.workbook.remove(self.workbook['Sheet'])
        # 保存工作簿
        self.workbook.save(output_address)
class Ruoguangchuli:
    def __init__(self):
        self.ruoguang_daysData= {}
        self.pianqumingxi=pd.DataFrame()
        self.chongfu_data_huizong_final=[]
        self.chongfu_data_final=[]
    # 存储数据界面
    def save_data_layout(self):
        layout = [
            [sg.Input(key='OUT_ADDRESS', disabled=True), sg.FileSaveAs(target='OUT_ADDRESS',file_types=(('All Files', '*.xlsx'),))],
            [sg.OK()]]
        window = sg.Window('保存倒出表格', layout)
        while True:
            event, values = window.read()
            if event == None:
                break
            elif event == 'OK':
                self.save_data(values['OUT_ADDRESS'])
                sg.Popup('文件生成成功！')
                break
        window.close()
    # 程序运行函数
    def run(self):
        layout = [
            [sg.Input(key='-PIANQU_IN-', disabled=True), sg.FileBrowse(button_text='片区明细',target='-PIANQU_IN-')],
            [sg.Input(key='-FILES_IN-', disabled=True), sg.FolderBrowse(button_text='5日内弱光文件夹',target='-FILES_IN-')],
            [sg.OK()]]
        window = sg.Window('弱光分析', layout)
        while True:
            event, values = window.read()
            if event == None:
                break
            # 运行功能方法
            if event == 'OK':
                self.get_data(values['-PIANQU_IN-'],values['-FILES_IN-'])
                self.solve_data()
                self.save_data_layout()
                break
    # 查找重复值函数
    @staticmethod
    def finddata(lst1,lst2):
        lst=[]
        for d in lst1:
            if d in lst2:
                lst.append('是')
            else:
                lst.append('否')
        return lst
    # 获取数据函数
    def get_data(self,pianqu_in,files_in):
        self.pianqumingxi=pd.read_excel(pianqu_in).loc[:,['小区名称','包户人姓名']]
        files = os.listdir(files_in)
        for file in files:
            riqi = pd.to_datetime(re.split('[._]', file)[1])
            df = pd.read_csv('弱光//' + file, encoding='gbk', encoding_errors='ignore')
            df = df[df['区县'] == '闻喜县']
            self.ruoguang_daysData[riqi] = df
    # 处理数据函数
    def solve_data(self):
        recent_day = max(self.ruoguang_daysData.keys())
        recent_data = self.ruoguang_daysData[recent_day]
        recent_data_weihu = pd.merge(recent_data, self.pianqumingxi, how='left', left_on='小区', right_on='小区名称')
        # 找出重复弱光的用户
        dic = {}
        for st in ['1 day', '2 day', '3 day', '4 day']:
            other = list(self.ruoguang_daysData[recent_day - pd.Timedelta(st)].loc[:, '用户地址'])
            shifouchongfu = self.finddata(list(recent_data_weihu['用户地址']), other)
            dic[st] = shifouchongfu
        chongfuruoguang = pd.DataFrame(dic)
        # 计算5天内共有几天弱光
        chongfuruoguang['5天内弱光天数'] = chongfuruoguang.apply(lambda s: str(sum(s == '是') + 1)+'天', axis=1)
        chongfu_data = pd.merge(recent_data_weihu, chongfuruoguang, left_index=True, right_index=True)
        chongfu_data_huizong = chongfu_data.loc[:, ['包户人姓名', '5天内弱光天数','SN']].pivot_table(
            index='包户人姓名',columns='5天内弱光天数', values='SN', aggfunc=np.count_nonzero,margins=True,margins_name='合计',fill_value=0)
        # 转换为列表，便于写入
        self.chongfu_data_huizong_final = np.vstack((chongfu_data_huizong.columns.insert(0, '处理人'), np.column_stack(
            (chongfu_data_huizong.index, chongfu_data_huizong.values)))).tolist()
        self.chongfu_data_final = [chongfu_data.columns.tolist()] + chongfu_data.values.tolist()
    # 保存数据函数
    def save_data(self,out_address):
        nwb = Workbook()
        nws1 = nwb.create_sheet('弱光明细')
        for row in self.chongfu_data_final:
            nws1.append(row)
        nws2 = nwb.create_sheet('汇总')
        for row in self.chongfu_data_huizong_final:
            nws2.append(row)
        # 设置网格线
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for row in nws2.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(name='宋体', size=12, bold=False, color='000000')
                    cell.border = thin_border
        nwb.remove(nwb['Sheet'])
        nwb.save(out_address)
class H5Tousu:
    def __init__(self):
        self.data_merge =pd.DataFrame()
        self.file_path = ''

        self.tousu_baobiao = pd.DataFrame()   # 储存投诉H5整体指标表
        self.huizong_by_renyuan = pd.DataFrame()  # 储存按维护员统计的投诉H5数量
        self.tousu_h5_detail = pd.DataFrame()   # 储存投诉H5明细
    def run(self):
        # 系统布局
        layout = [
            [sg.Text('请确保导入的表格中有‘系统’、‘值班’两个sheet',font=('宋体',16),text_color='black')],
            [sg.Input(key='-IN-', disabled=True), sg.FileBrowse(target='-IN-', file_types=(('All Files', '*.xlsx'),))],
            [sg.Input(key='-OUT-', disabled=True),
             sg.FileSaveAs(target='-OUT-', file_types=(('All Files', '*.xlsx'),))],
            [sg.OK()]]
        window = sg.Window('指标统计', layout)
        while True:
            event, values = window.read()
            if event == None:
                break
            # 运行功能方法
            if event == 'OK':
                self.merge_data(file_path=values['-IN-'])
                self.H5_detail()
                self.tousu_huizong()
                self.H5_tongbao()
                nwb = pd.ExcelWriter(values['-OUT-'])
                self.tousu_h5_detail.to_excel(excel_writer=nwb,
                                              sheet_name='投诉H5明细',
                                              index=False)
                self.tousu_baobiao.to_excel(excel_writer=nwb,
                                            sheet_name='投诉H5汇总',
                                            index=False)
                self.huizong_by_renyuan.to_excel(excel_writer=nwb,
                                                 sheet_name='维护员投诉H5统计',
                                                 index=False)
                nwb.close()
                wb1 = load_workbook(values['-OUT-'])
                wb1['投诉H5汇总'].insert_rows(0,1)
                wb1['投诉H5汇总']['a1'] = '投诉H5完成情况通报'
                wb1['投诉H5汇总'].merge_cells('a1:d1')
                set_excel_style(wb1,values['-OUT-'])
                sg.Popup('生成表格成功')
        window.close()
    def merge_data(self,file_path):
        self.file_path = file_path
        xitong_table = pd.read_excel(file_path, sheet_name='系统')
        zhiban_table = pd.read_excel(file_path, sheet_name='值班')
        self.data_merge = pd.merge(
            left=xitong_table, right=zhiban_table,
            left_on='开通工单号', right_on='工单流水号',
            how='left')
        self.data_merge['维护员'] = self.data_merge['维护员'].fillna('外县')
    def tousu_huizong(self):
        tousu_huizong = self.data_merge.groupby('维护员').agg(
            **{'下发量': ('开通工单号', 'count'), '回收量': ('评分', 'count')})
        tousu_huizong['回收率'] = (tousu_huizong['回收量'] / tousu_huizong['下发量']).apply(
            lambda x: '{:.2%}'.format(x))
        tousu_huizong.reset_index(inplace=True)
        tousu_zongji = pd.DataFrame(
            {'维护员': '合计', '下发量': tousu_huizong['下发量'].sum(),
             '回收量': tousu_huizong['回收量'].sum(),
             '回收率': '{:2%}'.format(tousu_huizong['回收量'].sum() / tousu_huizong['下发量'].sum())}, index=['合计'])
        self.huizong_by_renyuan = pd.concat(objs=[tousu_huizong, tousu_zongji], axis=0)

    def H5_tongbao(self):
        weixiu_h5_xiafa = self.data_merge['开通工单号'].count()
        weixiu_h5_huishou = self.data_merge['评分'].count()
        weixiu_huishou_rate = '{:.2%}'.format(weixiu_h5_huishou / weixiu_h5_xiafa)
        weixiu_manyi_rate = (self.data_merge['评分'].mean() - 1) / 9 * 100
        self.tousu_baobiao = pd.DataFrame(
            {'下发量': weixiu_h5_xiafa, '回收量': weixiu_h5_huishou, '回收率': weixiu_huishou_rate,
             '满意率': weixiu_manyi_rate}, index=[0])
    def H5_detail(self):
        self.tousu_h5_detail = self.data_merge.loc[:,
                          ['开通工单号', '用户电话', '维护员', '地址', '小区', '回单时间', '调研短信触发时间', '评分']]
        self.tousu_h5_detail['是否回收'] = self.data_merge['评分'].apply(lambda x: '是' if x / x == 1 else '否')

class H5Anzhuang:
    def __init__(self):
        self.data_merge = pd.DataFrame()
        self.file_path = ''

        self.anzhuang_baobiao = pd.DataFrame()  # 储存安装H5整体指标表
        self.huizong_by_renyuan = pd.DataFrame()  # 储存按维护员统计的安装H5数量
        self.anzhuang_h5_detail = pd.DataFrame()  # 储存安装H5明细
    def run(self):
        # 系统布局
        layout = [
            [sg.Text('请确保导入表格中有‘安装’、‘在跑’两个sheet',font=('宋体', 16), text_color='black')],
            [sg.Input(key='-IN-', disabled=True), sg.FileBrowse(target='-IN-', file_types=(('All Files', '*.xlsx'),))],
            [sg.Input(key='-OUT-', disabled=True),
             sg.FileSaveAs(target='-OUT-', file_types=(('All Files', '*.xlsx'),))],
            [sg.OK()]]
        window = sg.Window('指标统计', layout)
        while True:
            event, values = window.read()
            if event == None:
                break
            # 运行功能方法
            if event == 'OK':
                self.merge_data(file_path=values['-IN-'])
                self.H5_detail()
                self.tousu_huizong()
                self.H5_renyuan_huizong()
                nwb = pd.ExcelWriter(values['-OUT-'])
                self.anzhuang_h5_detail.to_excel(excel_writer=nwb,
                                                 sheet_name='安装H5明细',
                                                 index=False)
                self.anzhuang_baobiao.to_excel(excel_writer=nwb,
                                               sheet_name='安装H5汇总',
                                               index=False)
                self.huizong_by_renyuan.to_excel(excel_writer=nwb,
                                                 sheet_name='维护员H5统计',
                                                 index=False)
                nwb.close()
                wb1 = load_workbook(values['-OUT-'])
                wb1['安装H5汇总'].insert_rows(0,1)
                wb1['安装H5汇总']['a1'] = '安装H5完成情况通报'
                wb1['安装H5汇总'].merge_cells('a1:d1')
                set_excel_style(wb1,values['-OUT-'])

                sg.Popup('生成表格成功')
        window.close()
    def merge_data(self,file_path):
        self.file_path = file_path
        anzhuang_table = pd.read_excel(file_path,sheet_name='安装')
        zaipao_table = pd.read_excel(file_path, sheet_name='在跑')
        self.data_merge = pd.merge(
            left=anzhuang_table, right=zaipao_table,
            left_on='开通工单号', right_on='CRM订单号',
            how='left')
        self.data_merge['装机人员'] = self.data_merge['装机人员'].fillna('外县')
    def tousu_huizong(self):
        anzhuang_h5_xiafa = self.data_merge['开通工单号'].count()
        anzhuang_h5_huishou = self.data_merge['家宽装机打分'].count()
        anzhuang_huishou_rate = '{:.2%}'.format(anzhuang_h5_huishou / anzhuang_h5_xiafa)
        anzhuang_manyi_rate = (self.data_merge['家宽装机打分'].mean() - 1) / 9 * 100
        self.anzhuang_baobiao = pd.DataFrame(
            {'下发量': anzhuang_h5_xiafa, '回收量': anzhuang_h5_huishou, '回收率': anzhuang_huishou_rate,
             '满意率': anzhuang_manyi_rate}, index=[0])

    def H5_renyuan_huizong(self):
        huizong_by_renyuan = self.data_merge.groupby('装机人员').agg(
            **{'下发量': ('开通工单号', 'count'), '回收量': ('家宽装机打分', 'count')})
        huizong_by_renyuan['回收率'] = (huizong_by_renyuan['回收量'] / huizong_by_renyuan['下发量']).apply(
            lambda x: '{:.2%}'.format(x))
        huizong_by_renyuan.reset_index(inplace=True)
        tousu_zongji = pd.DataFrame(
            {'装机人员': '合计', '下发量': huizong_by_renyuan['下发量'].sum(),
             '回收量': huizong_by_renyuan['回收量'].sum(),
             '回收率': '{:2%}'.format(huizong_by_renyuan['回收量'].sum() / huizong_by_renyuan['下发量'].sum())}, index=['合计'])
        self.huizong_by_renyuan = pd.concat(objs=[huizong_by_renyuan, tousu_zongji], axis=0)


    def H5_detail(self):
        self.anzhuang_h5_detail = self.data_merge.loc[:,
                             ['开通工单号', '用户联系电话', '装机人员', '小区名称', '派单时间', '归档时间',
                              '调研短信触发时间', '调研短信回复时间']]
        self.anzhuang_h5_detail['是否回收'] = self.data_merge['家宽装机打分'].apply(lambda x: '是' if x / x == 1 else '否')


if __name__=='__main__':
    z=Zhuangjizhibiao()
    r=Ruoguangchuli()
    az=H5Anzhuang()
    ts = H5Tousu()
    layout = [
        [sg.B('装机',size=(20,3),font=('黑体',23)),sg.B('弱光',size=(20,3),font=('黑体',23)) ],
        [sg.B('投诉H5', size=(20, 3), font=('黑体', 23)), sg.B('安装H5', size=(20, 3), font=('黑体', 23))]]
    window = sg.Window('指标统计', layout)
    while True:
        event, values = window.read()
        if event == None:
            break
        elif event == '装机':
            z.run()
        elif event == '弱光':
            r.run()
        elif event == '投诉H5':
            ts.run()
        elif event == '安装H5':
            az.run()
    window.close()
